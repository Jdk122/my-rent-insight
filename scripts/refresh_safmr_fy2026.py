#!/usr/bin/env python3
"""
RenewalReply — Full National FMR Data Refresh (FY2026)
Two-pass load: SAFMR (ZIP-level) + ERAP (county-level fallback for missing ZIPs).

Usage:
  pip install openpyxl
  python scripts/refresh_safmr_fy2026.py

Required files in scripts/ directory:
  - fy2026_safmrs.xlsx     (from huduser.gov SAFMR page, ~4.3 MB)
  - fy2026_erap_fmrs.xlsx  (from huduser.gov FMR page, ~1.9 MB)

What it does:
  1. Pass 1: Reads fy2026_safmrs.xlsx — ZIP-level SAFMR data (~38K metro ZIPs)
  2. Pass 2: Reads fy2026_erap_fmrs.xlsx — fills remaining ZIPs with county-level FMR
  3. For existing ZIPs: moves current f → p (prior year), updates FMR, recalculates YoY
  4. For new ZIPs: adds them with FMR data and inferred metro/state
  5. NEVER deletes existing ZIPs — only adds and updates
  6. Writes updated rentData.json
"""

import json, os, sys

try:
    import openpyxl
except ImportError:
    print("Install openpyxl first:  pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SAFMR_PATH = os.path.join(SCRIPT_DIR, "fy2026_safmrs.xlsx")
ERAP_PATH = os.path.join(SCRIPT_DIR, "fy2026_erap_fmrs.xlsx")
RENT_DATA_PATH = os.path.join(SCRIPT_DIR, "..", "public", "data", "rentData.json")


def find_col(header_lower, *candidates):
    """Find column index by matching header text (case-insensitive, substring match)."""
    for c in candidates:
        c_lower = c.lower()
        for i, h in enumerate(header_lower):
            if c_lower in h:
                return i
    return None


def parse_safmr(path):
    """Parse SAFMR Excel file — returns dict of zip → {fmr: [...], metro: str}"""
    print(f"Pass 1: Parsing SAFMR file ({path})...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    header_lower = [str(h).strip().lower() if h else "" for h in header]

    col_zip = find_col(header_lower, "zip", "zcta", "zip_code", "zipcode")
    col_metro = find_col(header_lower, "hud fair market rent area name", "metro_name", "hmfa_name", "areaname", "area_name", "safmr_area_name")
    col_0br = find_col(header_lower, "safmr\n0br", "safmr 0br", "safmr_0br", "safmr0br")
    col_1br = find_col(header_lower, "safmr\n1br", "safmr 1br", "safmr_1br", "safmr1br")
    col_2br = find_col(header_lower, "safmr\n2br", "safmr 2br", "safmr_2br", "safmr2br")
    col_3br = find_col(header_lower, "safmr\n3br", "safmr 3br", "safmr_3br", "safmr3br")
    col_4br = find_col(header_lower, "safmr\n4br", "safmr 4br", "safmr_4br", "safmr4br")

    if col_zip is None or col_0br is None:
        print("  Could not find required SAFMR columns. Headers:")
        for i, h in enumerate(header):
            print(f"    [{i}] {repr(h)}")
        sys.exit(1)

    print(f"  ZIP col: {col_zip}, Metro col: {col_metro}")
    print(f"  FMR cols: 0BR={col_0br}, 1BR={col_1br}, 2BR={col_2br}, 3BR={col_3br}, 4BR={col_4br}")

    data = {}
    for row in rows:
        z = str(row[col_zip] or "").strip().zfill(5)
        if len(z) != 5 or not z.isdigit():
            continue
        fmr = [
            int(row[col_0br] or 0),
            int(row[col_1br] or 0),
            int(row[col_2br] or 0),
            int(row[col_3br] or 0),
            int(row[col_4br] or 0),
        ]
        if all(v == 0 for v in fmr):
            continue
        metro = str(row[col_metro] or "").strip() if col_metro is not None else ""
        data[z] = {"fmr": fmr, "metro": metro, "source": "safmr"}

    wb.close()
    print(f"  Parsed {len(data)} ZIP codes from SAFMR file.")
    return data


def parse_erap(path, exclude_zips):
    """Parse ERAP FMR file — returns dict of zip → {fmr: [...], metro: str} for ZIPs NOT in exclude_zips."""
    print(f"\nPass 2: Parsing ERAP file ({path})...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    header_lower = [str(h).strip().lower() if h else "" for h in header]

    # ERAP files typically use column names like: zcta, area_name, erap_0br, erap_1br, etc.
    # or: zip_code, fmr_area_name, Rent_0BR, Rent_1BR, etc.
    col_zip = find_col(header_lower, "zip", "zcta", "zip_code", "zipcode")
    col_metro = find_col(header_lower, "area_name", "metro_name", "hmfa_name", "areaname", "fmr_area_name", "safmr_area_name", "hud fair market rent area name")

    # Try various FMR column name patterns
    col_0br = find_col(header_lower, "erap_0br", "erap 0br", "rent_0br", "0br", "eff")
    col_1br = find_col(header_lower, "erap_1br", "erap 1br", "rent_1br", "1br")
    col_2br = find_col(header_lower, "erap_2br", "erap 2br", "rent_2br", "2br")
    col_3br = find_col(header_lower, "erap_3br", "erap 3br", "rent_3br", "3br")
    col_4br = find_col(header_lower, "erap_4br", "erap 4br", "rent_4br", "4br")

    if col_zip is None:
        print("  Could not find ZIP column in ERAP file. Headers:")
        for i, h in enumerate(header):
            print(f"    [{i}] {repr(h)}")
        print("  Skipping ERAP pass — only SAFMR ZIPs will be used.")
        wb.close()
        return {}

    if col_0br is None:
        # Try SAFMR-style columns as fallback
        col_0br = find_col(header_lower, "safmr_0br", "safmr0br", "safmr\n0br", "fmr_0")
        col_1br = find_col(header_lower, "safmr_1br", "safmr1br", "safmr\n1br", "fmr_1")
        col_2br = find_col(header_lower, "safmr_2br", "safmr2br", "safmr\n2br", "fmr_2")
        col_3br = find_col(header_lower, "safmr_3br", "safmr3br", "safmr\n3br", "fmr_3")
        col_4br = find_col(header_lower, "safmr_4br", "safmr4br", "safmr\n4br", "fmr_4")

    if col_0br is None:
        print("  Could not find FMR columns in ERAP file. Headers:")
        for i, h in enumerate(header):
            print(f"    [{i}] {repr(h)}")
        print("  Skipping ERAP pass.")
        wb.close()
        return {}

    print(f"  ZIP col: {col_zip}, Metro col: {col_metro}")
    print(f"  FMR cols: 0BR={col_0br}, 1BR={col_1br}, 2BR={col_2br}, 3BR={col_3br}, 4BR={col_4br}")

    data = {}
    total_in_file = 0
    for row in rows:
        z = str(row[col_zip] or "").strip().zfill(5)
        if len(z) != 5 or not z.isdigit():
            continue
        total_in_file += 1
        if z in exclude_zips:
            continue  # SAFMR already has this ZIP — skip
        fmr = [
            int(row[col_0br] or 0),
            int(row[col_1br] or 0) if col_1br is not None else 0,
            int(row[col_2br] or 0) if col_2br is not None else 0,
            int(row[col_3br] or 0) if col_3br is not None else 0,
            int(row[col_4br] or 0) if col_4br is not None else 0,
        ]
        if all(v == 0 for v in fmr):
            continue
        metro = str(row[col_metro] or "").strip() if col_metro is not None else ""
        data[z] = {"fmr": fmr, "metro": metro, "source": "county"}

    wb.close()
    print(f"  Total ZIPs in ERAP file: {total_in_file}")
    print(f"  New ZIPs not in SAFMR: {len(data)}")
    return data


def infer_state_city(metro):
    """Try to extract state abbreviation from metro name like 'Abilene, TX MSA'."""
    state = ""
    city = ""
    if ", " in metro:
        parts = metro.rsplit(", ", 1)
        if len(parts) == 2:
            state_part = parts[1].split()[0] if parts[1] else ""
            if len(state_part) == 2 and state_part.isalpha():
                state = state_part.upper()
    return city, state


def main():
    # ── Validate input files ──
    if not os.path.exists(SAFMR_PATH):
        print(f"ERROR: {SAFMR_PATH} not found.")
        print("  Download from: https://www.huduser.gov/portal/datasets/fmr/smallarea/index.html")
        sys.exit(1)

    has_erap = os.path.exists(ERAP_PATH)
    if not has_erap:
        print(f"WARNING: {ERAP_PATH} not found. Only SAFMR ZIPs will be loaded.")
        print("  Download from: https://www.huduser.gov/portal/datasets/fmr.html")
        print("  (Look for 'FY 2026 ERAP FMRs')")

    # ── Pass 1: SAFMR (ZIP-level granularity) ──
    safmr_data = parse_safmr(SAFMR_PATH)

    # ── Pass 2: ERAP (county-level fallback) ──
    erap_data = {}
    if has_erap:
        erap_data = parse_erap(ERAP_PATH, exclude_zips=set(safmr_data.keys()))

    # Combine: SAFMR takes priority
    all_new_data = {**safmr_data, **erap_data}
    print(f"\n  Combined: {len(all_new_data)} total ZIPs ({len(safmr_data)} SAFMR + {len(erap_data)} county-level)")

    # ── Load existing rentData.json ──
    print(f"\nLoading {RENT_DATA_PATH}...")
    if os.path.exists(RENT_DATA_PATH):
        with open(RENT_DATA_PATH) as f:
            rd = json.load(f)
        print(f"  Existing dataset: {len(rd)} ZIP codes.")
    else:
        rd = {}
        print("  No existing file — starting fresh.")

    # ── Process updates ──
    updated = 0
    added = 0
    unchanged = 0
    preserved = 0

    existing_zips = set(rd.keys())
    new_zips = set(all_new_data.keys())

    # Update existing ZIPs that appear in new data
    for z in existing_zips & new_zips:
        entry = rd[z]
        nd = all_new_data[z]
        new_fmr = nd["fmr"]
        old_fmr = entry.get("f", [0, 0, 0, 0, 0])

        # Shift current f → p (prior year)
        entry["p"] = old_fmr

        if old_fmr == new_fmr:
            unchanged += 1
            entry["y"] = 0.0
        else:
            entry["f"] = new_fmr
            updated += 1

        # Recalculate YoY
        if old_fmr[1] and old_fmr[1] > 0:
            yoy = round(((new_fmr[1] - old_fmr[1]) / old_fmr[1]) * 100, 1)
            yoy = max(-30, min(30, yoy))  # cap at ±30%
            entry["y"] = yoy

        # Update metro name if available
        metro = nd["metro"]
        if metro:
            entry["m"] = metro

        # Set FMR source flag
        entry["fs"] = nd["source"]

    # Add brand-new ZIPs
    for z in new_zips - existing_zips:
        nd = all_new_data[z]
        city, state = infer_state_city(nd["metro"])
        rd[z] = {
            "c": city,
            "s": state,
            "m": nd["metro"],
            "f": nd["fmr"],
            "p": [0, 0, 0, 0, 0],
            "y": 0.0,
            "ps": "f",
            "fs": nd["source"],
        }
        added += 1

    # Preserve existing ZIPs NOT in new data (never delete)
    orphans = existing_zips - new_zips
    preserved = len(orphans)
    if preserved > 0:
        print(f"\n  ⚠ {preserved} ZIPs in existing data not found in new source — PRESERVED (not deleted).")
        for z in sorted(orphans)[:10]:
            entry = rd[z]
            print(f"    {z}: {entry.get('c', '?')}, {entry.get('s', '?')}")
        if preserved > 10:
            print(f"    ... and {preserved - 10} more")

    # ── Report ──
    safmr_count = sum(1 for z in rd if rd[z].get("fs") == "safmr")
    county_count = sum(1 for z in rd if rd[z].get("fs") == "county")
    no_fs = sum(1 for z in rd if "fs" not in rd[z])

    print(f"\n  Results:")
    print(f"    Updated FMR: {updated} ZIPs")
    print(f"    Unchanged FMR: {unchanged} ZIPs")
    print(f"    New ZIPs added: {added}")
    print(f"    Preserved (not in source): {preserved} ZIPs")
    print(f"    Total ZIPs in output: {len(rd)}")
    print(f"      └─ SAFMR (ZIP-level): {safmr_count}")
    print(f"      └─ County-level FMR: {county_count}")
    if no_fs > 0:
        print(f"      └─ No source flag (legacy): {no_fs}")

    # ── Show samples ──
    print(f"\n  Sample updates (first 5 changed):")
    count = 0
    for z in sorted(existing_zips & new_zips):
        if rd[z].get("f") != rd[z].get("p") and count < 5:
            print(f"    {z}: 1BR=${rd[z]['f'][1]} (was ${rd[z]['p'][1]}), YoY={rd[z].get('y', 0)}%, source={rd[z].get('fs', '?')}")
            count += 1

    if added > 0:
        print(f"\n  Sample new ZIPs (first 10):")
        for z in sorted(new_zips - existing_zips)[:10]:
            print(f"    {z}: 1BR=${rd[z]['f'][1]}, source={rd[z]['fs']}, metro={rd[z].get('m', '?')}")

    # ── Confirm and write ──
    confirm = input(f"\n  Write updated rentData.json with {len(rd)} ZIPs? (y/N): ").strip().lower()
    if confirm != "y":
        print("  Aborted.")
        return

    with open(RENT_DATA_PATH, "w") as f:
        json.dump(rd, f, separators=(",", ":"))

    size_mb = os.path.getsize(RENT_DATA_PATH) / 1024 / 1024
    print(f"  Wrote rentData.json ({size_mb:.1f} MB) with {len(rd)} ZIPs. Done!")
    print(f"\n  Next steps:")
    print(f"    1. Run refresh_zori.py to add Zillow ZORI data")
    print(f"    2. Run refresh_zhvi.py to add Zillow ZHVI data (when available)")
    print(f"    3. Run refresh_apartmentlist.py to add Apartment List data (when available)")
    print(f"    4. Run refresh_hud50.py to add HUD 50th percentile data (when available)")
    print(f"    5. Update public/data/data_freshness.json with today's date")
    print(f"    6. Check /admin/data-quality for anomalies")


if __name__ == "__main__":
    main()
